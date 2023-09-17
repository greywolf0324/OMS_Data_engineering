import pdfplumber
import re
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
import os

class BUCEE_Parsing:
    def __init__(self) -> None:
        self.keys = ['LINE', 'SKU', 'VENDOR PN', 'UPC/GTIN', 'DESCRIPTIONLINE ITEM COMMENTS', 'MARKS AND NUMBERS', 'UNIT COST/RETAIL PRICE', 'QTY', 'UOM', 'ITEMTOTAL', 'PO Date:', 'Requested Delivery Date:', 'Requested Ship Date:', 'Cancel Date:', 'Delivery Window:', 'Shipping Window:', 'Vendor #:', 'Department #:', 'Freight Terms:', 'Preferred Carrier:', 'Terms Type', 'Terms Basis:', 'Terms Disc\n%:', 'Disc. Due Date:', 'Disc. Days:', 'Net Due Date:', 'Net Days:', 'Description:', 'TYPE', 'SERVICE TYPE', 'PERCENT', 'RATE', 'QTY_', 'UOM_', 'DESCRIPTION', 'AMOUNT', 'Total Qty:', 'Weight:', 'Volume:', 'Purchase Order Total:', '280.80', 'Order #', "PO_currency", "Ship To Name", "Ship To Address 1", "Ship To City", "Ship To State", "Ship to Zip", "Ship To Country", "Buying Party Name"]


    def re_fun(self, str_input: str) -> dict:
        if str_input == None:
            return None
        loc = str_input.rfind(":")
        left, right = str_input[:loc] + str_input[loc], (str_input[loc:])[1:]
        
        return {left: right}
    
    def PO_parser(self, paths: list):
        #this function will generate PO table
        res = {}
        
        for k, path in enumerate(paths):
            res[f"PDF{k}"] = {}
            pdf = pdfplumber.open(path)
            
            # for page_num, page in enumerate(pdf.pages):
            # # page = pdf.pages[1]
            #     res[f"PDF{k}"][f"page{page_num}"] = []
            #     cols_add = []
            #     for i, ele in enumerate(page.extract_tables()):
            #         if i != 2:
            #             items = [item for sublist in ele for item in sublist]
            #             for item in items: 
            #                 if self.re_fun(item) == None:
            #                     continue
            #                 else:
            #                     cols_add.append(self.re_fun(item))
            #     table_main = page.extract_tables()[2]
                
            #     for i, item in enumerate(table_main):
            #         for dic in cols_add:
            #             if i == 0: item.append(list(dic.keys())[0])
            #             elif i == 1: item.append(list(dic.values())[0].replace("\n", ""))
            #             else: item.append("")
            #     res[f"PDF{k}"][f"page{page_num}"].append(table_main)

            for page_num, page in enumerate(pdf.pages):
                res[f"PDF{k}"][f"page{page_num}"] = {}

                for key in self.keys:
                    res[f"PDF{k}"][f"page{page_num}"].update({key: []})
                
                #non main table addition
                non_product_tables = []
                ## non table addition
                non_table_content = []
                non_table_str = page.extract_text_simple()
                non_table_lis = non_table_str.split("\n")
                non_product_tables.append({"Order #": non_table_lis[4]})
                non_product_tables.append({"Ship To Name": "Buc" + non_table_lis[18].split("Buc")[1]})
                non_product_tables.append({"Ship To Address 1": non_table_lis[19]})
                non_product_tables.append({"Ship To City": non_table_lis[20].split(",")[0]})
                non_product_tables.append({"Ship To State": non_table_lis[20].split(",")[1].split(" ")[1]})
                non_product_tables.append({"Ship to Zip": non_table_lis[20].split(",")[1].split(" ")[2]})
                non_product_tables.append({"Ship To Country": non_table_lis[20].split(",")[1].split(" ")[3]})
                non_product_tables.append({"Buying Party Name":("Buc" + non_table_lis[18].split("Buc")[2]).split("Creative ")[0]})

                for i, table in enumerate(page.extract_tables()):
                    if i == 2:
                        product_table = table
                    if i != 2:
                        items = [item for sublist in table for item in sublist]
                        for item in items: 
                            if self.re_fun(item) == None:
                                continue
                            else:
                                non_product_tables.append(self.re_fun(item))

                non_product_dic = {}
                for dic in non_product_tables:
                    non_product_dic.update(dic)
                
                non_blank_keys = list(non_product_dic.keys())

                for key in self.keys:
                    if key == "UOM_":
                        res[f"PDF{k}"][f"page{page_num}"][key].append(non_product_dic["UOM"])
                    elif key == "QTY_":
                        res[f"PDF{k}"][f"page{page_num}"][key].append(non_product_dic["QTY"])
                    elif key == "PO_currency":
                        res[f"PDF{k}"][f"page{page_num}"][key].append("USD")
                    else:
                        if key in non_blank_keys:
                            res[f"PDF{k}"][f"page{page_num}"][key].append(non_product_dic[key])
                        else:
                            res[f"PDF{k}"][f"page{page_num}"][key].append("")

                # main table addition
                
                # product_keys = product_table[0]
                product_table_ = product_table[1 : len(product_table) - 1]
                product_keys = ['LINE', 'SKU', 'VENDOR PN', 'UPC/GTIN', 'DESCRIPTIONLINE ITEM COMMENTS', 'MARKS AND NUMBERS', 'UNIT COST/RETAIL PRICE', 'QTY', 'UOM', 'ITEMTOTAL']
                product_dic = {}
                for key in product_keys:
                    product_dic.update({key:[]})

                for i, line in enumerate(product_table_):
                    for i, key in enumerate(product_keys):
                        product_dic[key].append(line[i])
                for i in range(len(product_table_)):
                    for key in self.keys:
                        if key in product_keys:
                            res[f"PDF{k}"][f"page{page_num}"][key].append(product_dic[key][i])
                        else:
                            res[f"PDF{k}"][f"page{page_num}"][key].append("")
                temp_total = [product_table[len(product_table) - 1][9]]
                for _ in range(len(product_table) - 2):
                    temp_total.append("")
                res[f"PDF{k}"][f"page{page_num}"].update({"total": temp_total})

        if os.path.isfile("OCR_res.xlsx"):
            os.remove("OCR_res.xlsx")
        book = xlsxwriter.Workbook("OCR_res.xlsx")
        sheet = book.add_worksheet("cont_excel")
        for idx, header in enumerate(self.keys):
            sheet.write(0, idx, header)
        sheet.write(0, len(self.keys), "total")
        book.close()

        book = load_workbook("OCR_res.xlsx")
        sheet = book.get_sheet_by_name("cont_excel")
        for dic in res:
            for di in res[dic]:
                for i in range(len(res[dic][di]["LINE"])):
                    temp = []

                    for key in res[dic][di].keys():
                        temp.append(res[dic][di][key][i])

                    sheet.append(temp)


        book.save(filename = "OCR_res.xlsx")
        
        # pd.DataFrame(data = temp).to_excel("temp.xlsx", index = False)
        return res
    
# PDF Parsing for PEPCO
class PEPCO_Parsing:
    def __init__(self) -> None:
        self.keys = ["Order - ID", "Pre Order -ID", "Item No", "Item classification", "Item name", "Item name English", "Promotional product", "Supplier product code", "Season", "Merch code", "Collection", "Pictogram no", "Style type", "Supplier name", "Supplier ID", "Terms of payments", "Date of order creation", "Booking date", "Handover date", "Port of shipment", "Destination port", "Destination DC", "Delivery terms", "Transport mode", "Time of delivery", "Purchase price", "Total", "ONE", "Pack multiplier", "Total qty in outer", "barcode", "PO_currency"]
        pass

    def PO_parser(self, paths: list, currency):
        #this function will generate PO table
        res = {}

        for k, path in enumerate(paths):
            res[f"PDF{k}"] = {}
            pdf = pdfplumber.open(path)

            for key in self.keys:
                res[f"PDF{k}"].update({key: []})

            for page_num, page in enumerate(pdf.pages):
                if page_num == 0:
                    

                    content = page.extract_text_simple().split("\n")
                    if currency == "eur":
                        #title - 1
                        for i in range(13):
                            # res[f"PDF{k}"].update(
                            #     {
                            #         self.keys[i]: content[i + 5].split(".")[-1][1:]
                            #     }
                            # )
                            res[f"PDF{k}"][self.keys[i]].append(content[i + 5].split(".")[-1][1:])

                        res[f"PDF{k}"][self.keys[0]].append(content[0 + 5].split(".")[-1][1:])
                        
                        for i in range(1, 13):
                            res[f"PDF{k}"][self.keys[i]].insert(0, "")

                        #title - 2
                        for i in range(4):
                            # res[f"PDF{k}"].update(
                            #     {
                            #         self.keys[i + 13]: content[i + 19].split(".")[-1][1:]
                            #     }
                            # )
                            res[f"PDF{k}"][self.keys[i + 13]].append(content[i + 19].split(".")[-1][1:])

                        for i in range(3):
                            res[f"PDF{k}"][self.keys[i + 13]].insert(0, "")
                        res[f"PDF{k}"][self.keys[3 + 13]].insert(1, "")

                        #title - 3
                        for i in range(8):
                            # res[f"PDF{k}"].update(
                            #     {
                            #         self.keys[i + 17]: content[i + 24].split(".")[-1][1:]
                            #     }
                            # )
                            res[f"PDF{k}"][self.keys[i + 17]].append(content[i + 24].split(".")[-1][1:])
                        
                        for i in range(8):
                            if i == 0:
                                res[f"PDF{k}"][self.keys[i + 17]].insert(1, "")
                            elif i == 1:
                                res[f"PDF{k}"][self.keys[i + 17]].insert(1, "")
                            else:
                                res[f"PDF{k}"][self.keys[i + 17]].insert(0, "")

                    if currency == "usd":
                        #title - 1
                        for i in range(13):
                            # res[f"PDF{k}"].update(
                            #     {
                            #         self.keys[i]: content[i + 4].split(".")[-1][1:]
                            #     }
                            # )
                            res[f"PDF{k}"][self.keys[i]].append(content[i + 4].split(".")[-1][1:])
                        res[f"PDF{k}"][self.keys[0]].append(content[0 + 4].split(".")[-1][1:])

                        for i in range(1, 13):
                            res[f"PDF{k}"][self.keys[i]].insert(0, "")

                        #title - 2
                        for i in range(4):
                            # res[f"PDF{k}"].update(
                            #     {
                            #         self.keys[i + 13]: content[i + 18].split(".")[-1][1:]
                            #     }
                            # )
                            res[f"PDF{k}"][self.keys[i + 13]].append(content[i + 18].split(".")[-1][1:])

                        for i in range(3):
                            res[f"PDF{k}"][self.keys[i + 13]].insert(0, "")
                        res[f"PDF{k}"][self.keys[3 + 13]].insert(1, "")

                        #title - 3
                        for i in range(7):
                            # res[f"PDF{k}"].update(
                            #     {
                            #         self.keys[i + 17]: content[i + 23].split(":")[-1]
                            #     }
                            # )
                            res[f"PDF{k}"][self.keys[i + 17]].append(content[i + 23].split(":")[-1])
                        # res[f"PDF{k}"].update(
                        #         {
                        #             self.keys[24]: ":".join(content[30].split(":")[-3:])
                        #         }
                        #     )
                        res[f"PDF{k}"][self.keys[24]].append(":".join(content[30].split(":")[-3:]))

                        for i in range(7):
                            if i == 0:
                                res[f"PDF{k}"][self.keys[i + 17]].insert(1, "")
                            elif i == 1:
                                res[f"PDF{k}"][self.keys[i + 17]].insert(1, "")
                            else:
                                res[f"PDF{k}"][self.keys[i + 17]].insert(0, "")
                        
                        res[f"PDF{k}"][self.keys[24]].insert(0, "")

                if page_num == 1:
                    content = page.extract_text_simple().split("\n")
                    if currency == "eur":
                        # res[f"PDF{k}"].update(
                        #         {
                        #             self.keys[25]: " ".join(content[6].split(" ")[3:])
                        #         }
                        #     )
                        res[f"PDF{k}"][self.keys[25]].append(" ".join(content[6].split(" ")[3]))
                        res[f"PDF{k}"][self.keys[25]].insert(1, "")
                        # res[f"PDF{k}"].update(
                        #         {
                        #             self.keys[26]: content[8].split(".")[-1][1:]
                        #         }
                        #     )
                        res[f"PDF{k}"][self.keys[26]].append(content[8].split(".")[-1][1:])
                        res[f"PDF{k}"][self.keys[26]].insert(0, "")

                        res[f"PDF{k}"]["PO_currency"].append(content[6].split(" ")[4])
                        res[f"PDF{k}"]["PO_currency"].append("")
                    if currency == "usd":
                        # res[f"PDF{k}"].update(
                        #         {
                        #             self.keys[25]: " ".join(content[5].split(" ")[2].split("\xa0")[0])
                        #         }
                        #     )
                        res[f"PDF{k}"][self.keys[25]].append(" ".join(content[5].split(" ")[2].split("\xa0")[0]))
                        res[f"PDF{k}"][self.keys[25]].insert(1, "")
                        # res[f"PDF{k}"].update(
                        #         {
                        #             self.keys[26]: content[7].split(".")[-1].replace("\xa0", "").replace(": ", "")
                        #         }
                        #     )
                        res[f"PDF{k}"][self.keys[26]].append(content[7].split(".")[-1].replace("\xa0", "").replace(": ", ""))
                        res[f"PDF{k}"][self.keys[26]].insert(0, "")

                        res[f"PDF{k}"]["PO_currency"].append(content[5].split(" ")[2].split("\xa0")[1])
                        res[f"PDF{k}"]["PO_currency"].append("")
                if page_num == 2:
                    content = page.extract_text_simple().split("\n")
                    if currency == "eur":
                        # res[f"PDF{k}"].update(
                        #         {
                        #             self.keys[27]: content[13].split(" ")[1]
                        #         }
                        #     )
                        res[f"PDF{k}"][self.keys[27]].append(content[13].split(" ")[1])
                        res[f"PDF{k}"][self.keys[27]].insert(0, "")
                        # res[f"PDF{k}"].update(
                        #         {
                        #             self.keys[28]: content[13].split(" ")[3]
                        #         }
                        #     )
                        res[f"PDF{k}"][self.keys[28]].append(content[13].split(" ")[3])
                        res[f"PDF{k}"][self.keys[28]].insert(0, "")
                        # res[f"PDF{k}"].update(
                        #         {
                        #             self.keys[29]: content[13].split(" ")[4]
                        #         }
                        #     )
                        res[f"PDF{k}"][self.keys[29]].append(content[13].split(" ")[4])
                        res[f"PDF{k}"][self.keys[29]].insert(0, "")

                        res[f"PDF{k}"][self.keys[30]].append(content[6].split(";")[1].split(":")[1][1:])
                        res[f"PDF{k}"][self.keys[30]].insert(0, "")
                    #USD
                    if currency == "usd":
                        # res[f"PDF{k}"].update(
                        #         {
                        #             self.keys[27]: content[10].split(" ")[1]
                        #         }
                        #     )
                        res[f"PDF{k}"][self.keys[27]].append(content[10].split(" ")[1])
                        res[f"PDF{k}"][self.keys[27]].insert(0, "")
                        # res[f"PDF{k}"].update(
                        #         {
                        #             self.keys[28]: content[10].split(" ")[3]
                        #         }
                        #     )
                        res[f"PDF{k}"][self.keys[28]].append(content[10].split(" ")[3])
                        res[f"PDF{k}"][self.keys[28]].insert(0, "")
                        # res[f"PDF{k}"].update(
                        #         {
                        #             self.keys[29]: content[10].split(" ")[4]
                        #         }
                        #     )
                        res[f"PDF{k}"][self.keys[29]].append(content[10].split(" ")[4])
                        res[f"PDF{k}"][self.keys[29]].insert(0, "")
                        
                        res[f"PDF{k}"][self.keys[30]].append(content[5].split(";")[1].split(":")[1].split("\xa0")[1])
                        res[f"PDF{k}"][self.keys[30]].insert(0, "")
            
            

        if os.path.isfile("OCR_res.xlsx"):
            os.remove("OCR_res.xlsx")
        book = xlsxwriter.Workbook("OCR_res.xlsx")
        sheet = book.add_worksheet("cont_excel")
        for idx, header in enumerate(self.keys):
            sheet.write(0, idx, header)
        sheet.write(0, len(self.keys), "total")
        book.close()

        book = load_workbook("OCR_res.xlsx")
        sheet = book.get_sheet_by_name("cont_excel")
        for dic in res:
            
            
            for i in range(2):
                temp = []
                for key in res[dic].keys():
                    temp.append(res[dic][key][i])

                sheet.append(temp)


        book.save(filename = "OCR_res.xlsx")
        return res