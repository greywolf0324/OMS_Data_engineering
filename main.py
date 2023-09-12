<<<<<<< HEAD
from OCR.parse import Buc_parsing, PEPCO_Parsing
from Data_Integration.matching import PO_Match
from DB_Updater.noticer import NOTICER
from DB_Updater.DB_Updater import Udpater
from Integration_Add.Integrator import BUC_Integrate_All, PEPCO_Integrate_All
=======
<<<<<<< HEAD
from OCR.parse import Buc_parsing, PEPCO_Parsing
=======
from OCR.parse import Buc_parsing
>>>>>>> a4aec287cb11ba11dc90c80e78685035af5176fb
from Data_Integration.matching import PO_Match
from DB_Updater.noticer import NOTICER
from DB_Updater.DB_Updater import Udpater
from Integration_Add.Integrator import Integrate_All
>>>>>>> 30d5ecb1a2a3f34d8b139374f3abfadbdafdd438
from OMS_Creation.oms import OMS_Creation
from csv import DictWriter
import json
import pandas as pd
import csv
from openpyxl import load_workbook
import xlsxwriter
import os

<<<<<<< HEAD
def Sys_Buc():
    paths = [r"E:\work\Daily\8_10\_N\OMS_Data_engineering\Experiment_result\input\PDF\Buc-EE's\sample.pdf"]
    
    # OCR : Parsing PDF and generate table results
    print("On PDF parsing...")
    parser = Buc_parsing()
=======
<<<<<<< HEAD
def Sys_Buc():
=======
def main(st: str):
>>>>>>> a4aec287cb11ba11dc90c80e78685035af5176fb
    paths = [r"E:\work\Daily\8_10\_N\OMS_Data_engineering\Experiment_result\input\PDF\Buc-EE's\multi.pdf"]
    
    # OCR : Parsing PDF and generate table results
    print("On PDF parsing...")
<<<<<<< HEAD
    parser = Buc_parsing()
=======
    parser = globals()[st]()
>>>>>>> a4aec287cb11ba11dc90c80e78685035af5176fb
>>>>>>> 30d5ecb1a2a3f34d8b139374f3abfadbdafdd438
    PO_res = parser.PO_parser(paths)

    # Data_Integration: Generate SalesImport_Original
    print("On Match Operating...")
    matcher = PO_Match()
    matching_res = matcher.match_final(PO_res)

    # # Notation: Add following Info to DB
    print("Getting newal things...")
    noticer = NOTICER()
    addition = noticer.getter(matching_res)
    print(addition)
    #get response from frontend based on "addition"
    response_from_frontend = [] #{"OMS_AdditionalUOM": "OMS_PaymentTerm": "OMS_InventoryList":}
    # ########################################################################################
    # ##    Add "addition" info to OMS_AdditionalUOM, OMS_PaymentTerm, OMS_InventoryList    ##
    # ########################################################################################
    # print("DB Updating...")
    # udpater = Udpater()
    # udpater.updater(response_from_frontend)

    # Integration_Add : Generate [SalesImport, new_sku, new_paymentterm]
    print("Integrating...")
<<<<<<< HEAD
    integreator = BUC_Integrate_All()
=======
    integreator = Integrate_All()
>>>>>>> 30d5ecb1a2a3f34d8b139374f3abfadbdafdd438
    SalesImport = integreator.Integrate_final(matching_res)
    
    # # Generating OMS
    # print("Generating OMS...")
    # generator = OMS_Creation()
    # generator.OMS_generator(SalesImport)
    
    print("Just a second, writing...")
    f = open("config/fieldnames_SalesImport.json")

    field_names = json.load(f)

    # generate excel output file
    if os.path.isfile("SalesImport.xlsx"):
        os.remove("SalesImport.xlsx")
    book = xlsxwriter.Workbook("SalesImport.xlsx")
    sheet = book.add_worksheet("cont_excel")
    keys = list(SalesImport[0].keys())
    for idx, header in enumerate(field_names):
        sheet.write(0, idx, header)

    book.close()

    book = load_workbook("SalesImport.xlsx")
    sheet = book.get_sheet_by_name("cont_excel")
    
    for dic in SalesImport:
        for i in range(len(dic[keys[0]])):
            temp = []

            for key in field_names:
                # print(len(res[dic][di]["LINE"]))
                # print(res[dic][di])
                # print(key)
                if key in keys:
                    temp.append(dic[key][i])
                else:
                    temp.append("")
            sheet.append(temp) 
    
    book.save(filename = "SalesImport.xlsx")

    print("successful!")

def Sys_PEPCO():
    paths = [r"E:\work\Daily\8_10\_N\OMS_Data_engineering\Experiment_result\input\PDF\PEPCO\ORD00913820_01 _ CK BRANDS LIMITED _ 264252_orderSupp.pdf"]
    # OCR : Parsing PDF and generate table results
    print("On PDF parsing...")
    parser = PEPCO_Parsing()
    PO_res = parser.PO_parser(paths)

<<<<<<< HEAD
    # Integration_Add: Generate SalesImport\
    print("Integrating...")
    integrator = PEPCO_Integrate_All()
    SalesImport = integrator.Integrate_final(PO_res)
    print(SalesImport)
    print("Just a second, writing...")
    f = open("config/fieldnames_SalesImport.json")

    field_names = json.load(f)

    # generate excel output file
    if os.path.isfile("SalesImport.xlsx"):
        os.remove("SalesImport.xlsx")
    book = xlsxwriter.Workbook("SalesImport.xlsx")
    sheet = book.add_worksheet("cont_excel")
    keys = list(SalesImport[0].keys())
    for idx, header in enumerate(field_names):
        sheet.write(0, idx, header)

    book.close()

    book = load_workbook("SalesImport.xlsx")
    sheet = book.get_sheet_by_name("cont_excel")
    
    for dic in SalesImport:
        for i in range(2):
            temp = []

            for key in field_names:
                # print(len(res[dic][di]["LINE"]))
                # print(res[dic][di])
                # print(key)
                if key in keys:
                    temp.append(dic[key][i])
                else:
                    temp.append("")
            sheet.append(temp) 
    
    book.save(filename = "SalesImport.xlsx")

    print("successful!")

if __name__ == "__main__":
    
    st = "Sys_PEPCO"
    
    globals()[st]()
=======
if __name__ == "__main__":
    
<<<<<<< HEAD
    st = "Sys_Buc"
    
    globals()[st]()
=======
    main("Buc_parsing")
>>>>>>> a4aec287cb11ba11dc90c80e78685035af5176fb
>>>>>>> 30d5ecb1a2a3f34d8b139374f3abfadbdafdd438
    
    