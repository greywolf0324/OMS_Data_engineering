import datetime
import uuid
from .OCR.parse import BUCEE_Parsing, PEPCO_Parsing
from .Data_Integration.matching import PO_Match_BUCEE, PO_Match_PEPCO
from .DB_Updater.equal_extractor import Extractor
from .DB_Updater.noticer import NOTICER
from .DB_Updater.DB_Updater import Udpater
from .Integration_Add.Integrator import Integrate_All

from .OMS_Creation.oms import OMS_Creation
from csv import DictWriter
import json
import pandas as pd
import csv
from openpyxl import load_workbook
import xlsxwriter
import os
from pathlib import Path

def uploadFile(data):
    current_datetime = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    random_string = str(uuid.uuid4().hex)
    filename = f'{current_datetime}_{random_string}'
    extension = data.name.split(".")[-1]
    path = f'process/inputs/{filename}.{extension}'
    print(data)
    with open(path, 'wb') as destination:
        for chunk in data.chunks():
            destination.write(chunk)
    return [path, filename]

def parseUpload(data):
    paths = []
    for file in data:
        res = uploadFile(file)
        paths.append(res[0])
    
    # OCR: PDF parsing
    print("==============================================================================================================")
    print("On PDF parsing...")
    parser = BUCEE_Parsing()
    PO_res = parser.PO_parser(paths)

    # Data_Integration : Generate SalesImport_Original
    print("==============================================================================================================")
    print("On Match Operating...")
    matcher = PO_Match_BUCEE()
    matching_res = matcher.match_final(PO_res)

    # # Notation: Add following Info to DB
    print("==============================================================================================================")
    print("Getting newal things for OMS Update...")
    
    # # Extract equal OMS
    extract = Extractor()
    OMS_equal = extract.extractor(matching_res)

    return [matching_res, OMS_equal]
        
def requiredFields(matching_res):
    noticer = NOTICER()
    addition = noticer.getter(matching_res)
    return addition

def processFile(data, matching_res, extra, customer_name):
    customer_name = "BUC-EE'S"
    paths = []
    for file in data:
        res = uploadFile(file)
        path = res[0]
        paths.append(res[0])
        filename = res[1]


    print("==============================================================================================================")
    print("DB Updating...")
    udpater = Udpater()
    udpater.updater(extra)

    # Integration_Add : Generate SalesImport
    print("==============================================================================================================")
    print("Integrating...")
    integreator = Integrate_All(customer_name=customer_name)
    sales_import = integreator.Integrate_final(matching_res, customer_name)
    
    # Generating OMS
    # print("Generating OMS...")
    # generator = OMS_Creation()
    # generator.OMS_generator(sales_import)
    print("==============================================================================================================")
    print("Just a second, writing...")
    f = open(Path(__file__).resolve().parent / "config/fieldnames_SalesImport.json")

    field_names = json.load(f)
    # with open('Exam/output/output.csv', 'w') as outfile:
    #     writer = DictWriter(outfile, field_names)
    #     writer.writeheader()
    #     writer.writerows(sales_import)


    # with open("output.json", 'w') as f:
    #     json.dump(sales_import, f)

    # generate excel output file
    if os.path.isfile("SalesImport.xlsx"):
        os.remove("SalesImport.xlsx")
    book = xlsxwriter.Workbook("SalesImport.xlsx")
    sheet = book.add_worksheet("cont_excel")
    keys = list(sales_import[0].keys())
    for idx, header in enumerate(field_names):
        sheet.write(0, idx, header)

    book.close()

    book = load_workbook("SalesImport.xlsx")
    sheet = book.get_sheet_by_name("cont_excel")
    
    for dic in sales_import:
        for i in range(len(dic[keys[0]])):
            temp = []

            for key in field_names:
                # print(len(res[dic][di]["LINE"]))
                # print(res[dic][di])
                # print(key)
                if key in keys:
                    temp.append(dic[key][i])
                else:
                    temp.append("")
            sheet.append(temp) 
    output = f'process/outputs/{filename}.xlsx'
    book.save(filename = output)

    return [path, output]

def parseUpload_2(data, currency):
    paths = []
    for file in data:
        res = uploadFile(file)
        paths.append(res[0])
    
    # OCR: PDF parsing
    print("==============================================================================================================")
    print("On PDF parsing...")
    parser = PEPCO_Parsing()
    PO_res = parser.PO_parser(paths, currency)

    # Data_Integration : Generate SalesImport_Original
    print("==============================================================================================================")
    print("On Match Operating...")
    matcher = PO_Match_PEPCO()
    matching_res = matcher.match_final(PO_res)

    # # Notation: Add following Info to DB
    print("==============================================================================================================")
    print("Getting newal things for OMS Update...")

    # # Extract equal OMS
    extract = Extractor()
    OMS_equal = extract.extractor(matching_res)

    return [matching_res, OMS_equal]
   
def processFile_2(data, matching_res, extra, currency):
    customer_name = "Pepco"
    paths = []
    for file in data:
        res = uploadFile(file)
        path = res[0]
        paths.append(res[0])
        filename = res[1]

    print("==============================================================================================================")
    print("DB Updating...")
    udpater = Udpater()
    udpater.updater(extra)

    # Integration_Add : Generate SalesImport
    print("Integrating...")
    integreator = Integrate_All(customer_name)
    sales_import = integreator.Integrate_final(matching_res, currency)
    
    print("Just a second, writing...")
    f = open(Path(__file__).resolve().parent / "config/fieldnames_SalesImport.json")

    field_names = json.load(f)

    # generate excel output file
    if os.path.isfile("SalesImport.xlsx"):
        os.remove("SalesImport.xlsx")
    book = xlsxwriter.Workbook("SalesImport.xlsx")
    sheet = book.add_worksheet("cont_excel")
    keys = list(sales_import[0].keys())
    for idx, header in enumerate(field_names):
        sheet.write(0, idx, header)

    book.close()

    book = load_workbook("SalesImport.xlsx")
    sheet = book.get_sheet_by_name("cont_excel")
    for dic in sales_import:
        for i in range(2):
            temp = []
            # print(keys, "++++++++++")
            for key in field_names:
                # print(len(res[dic][di]["LINE"]))
                # print(res[dic][di])
                # print(key)
                if key in keys:
                    temp.append(dic[key][i])
                else:
                    temp.append("")
            sheet.append(temp) 
    output = f'process/outputs/{filename}.xlsx'
    book.save(filename = output)

    return [path, output]